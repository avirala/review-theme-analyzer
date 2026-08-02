"""Builds the two-sheet Excel workbook: Theme Summary (live COUNTIFS/
AVERAGEIFS formulas) + Review Detail (one row per review)."""
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

FONT = "Arial"


def _theme_order(rows):
    """Order themes by frequency (General Positive/Negative and Other pinned
    to sensible positions), sub-themes by frequency within each theme."""
    counts = defaultdict(int)
    theme_totals = defaultdict(int)
    for r in rows:
        counts[(r["theme"], r["sub_theme"])] += 1
        theme_totals[r["theme"]] += 1

    themes_sorted = sorted(theme_totals, key=lambda t: -theme_totals[t])
    # keep "Other / Uncategorized" last regardless of raw frequency, matching
    # the convention used throughout the manual analyses
    if "Other / Uncategorized" in themes_sorted:
        themes_sorted.remove("Other / Uncategorized")
        themes_sorted.append("Other / Uncategorized")

    ordered_keys = []
    for theme in themes_sorted:
        subs = [(k, v) for k, v in counts.items() if k[0] == theme]
        subs.sort(key=lambda x: -x[1])
        ordered_keys.extend(k for k, v in subs)

    return themes_sorted, ordered_keys


def build_workbook(rows, output_path):
    themes_sorted, sheet1_keys = _theme_order(rows)

    wb = openpyxl.Workbook()
    header_font = Font(name=FONT, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    body_font = Font(name=FONT)

    ws1 = wb.active
    ws1.title = "Theme Summary"
    headers1 = ["#", "theme", "sub_theme", "count", "avg rating"]
    ws1.append(headers1)
    for c in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, (theme, sub_theme) in enumerate(sheet1_keys, start=1):
        row_idx = ws1.max_row + 1
        ws1.cell(row=row_idx, column=1, value=i).font = body_font
        ws1.cell(row=row_idx, column=2, value=theme).font = body_font
        ws1.cell(row=row_idx, column=3, value=sub_theme).font = body_font
        count_formula = (
            f"=COUNTIFS('Review Detail'!$A:$A,$B{row_idx},'Review Detail'!$B:$B,$C{row_idx})"
        )
        avg_formula = (
            f"=IFERROR(AVERAGEIFS('Review Detail'!$D:$D,'Review Detail'!$A:$A,$B{row_idx},"
            f"'Review Detail'!$B:$B,$C{row_idx}),0)"
        )
        ws1.cell(row=row_idx, column=4, value=count_formula).font = body_font
        avg_cell = ws1.cell(row=row_idx, column=5, value=avg_formula)
        avg_cell.font = body_font
        avg_cell.number_format = "0.00"

    ws1.freeze_panes = "A2"
    for i, w in enumerate([5, 55, 65, 10, 12], start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Review Detail")
    headers2 = ["theme", "sub_theme", "review", "rating", "timestamp"]
    ws2.append(headers2)
    for c in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    theme_rank = {t: i for i, t in enumerate(themes_sorted)}
    rows_sorted = sorted(rows, key=lambda r: (theme_rank[r["theme"]], r["sub_theme"]))
    for r in rows_sorted:
        row_idx = ws2.max_row + 1
        ws2.cell(row=row_idx, column=1, value=r["theme"]).font = body_font
        ws2.cell(row=row_idx, column=2, value=r["sub_theme"]).font = body_font
        review_cell = ws2.cell(row=row_idx, column=3, value=r["review_text"])
        review_cell.font = body_font
        review_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.cell(row=row_idx, column=4, value=int(r["rating"])).font = body_font
        date_val = r["date"]
        ws2.cell(row=row_idx, column=5, value=str(date_val) if date_val else "").font = body_font

    ws2.freeze_panes = "A2"
    for i, w in enumerate([55, 65, 90, 8, 20], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    return len(sheet1_keys), len(rows_sorted)
